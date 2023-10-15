"""
Este módulo é voltado para guardar o código do Web Scraping. 

Neste módulo é usando a classe Navegador, que herda as funcionalidades do Selenium.

O objetivo deste código é coletar os dados de figuras públicas através das redes sociais Facebook e Twitter.

Documentação Adicional: https://peps.python.org/pep-0008/ | https://selenium-python.readthedocs.io/index.html

Licença:
Este código está sujeito às políticas e regulamentos internos.

© 2023, Natan.
"""

# Importações do Python
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# Importações Internas
from engine.navegador import Navegador
from utilitarios.utilitarios import *

class Automacao:
    def raspar_facebook(self):
        """Função definida para raspar os dados de um perfil do Facebook."""

        # Salvando o input do usuário
        url = input("\033[34mInsira o nome do usuário ou cole o link do perfil: \033[0m")
        usuario_pesquisa = [i for i in url.split("/") if i.strip()][-1]

        # Verifica se deseja período
        filtrar_periodo = False
        inp = input("\033[34mDeseja informar período? [s/n]: \033[0m")
        if  inp.upper().strip() == "S":
            periodo_inicio = input("\033[34mInforme o início do período que deseja filtrar - [DD/MM/YYYY]: \033[0m")
            periodo_final = input("\033[34mInforme o final do período que deseja filtrar - [DD/MM/YYYY]: \033[0m")
            filtrar_periodo = True
        sleep(0.5)
        self.carregar()
        os.system('cls')
        self.carregar_iniciando()

        # Instanciando o Navegador
        self.navegador = Navegador(somente_terminal=True, salvar_cache=PATH['arquivo_cache'], tela_cheia=True)

        # Validar path da url
        if "?" in url:
            url = url.split("?")[0]

        # Verficia de o Input foi uma URL ou o nome do Usuário
        if self.e_url(url):
            if "mbasic" not in url:
                url = url.replace("web.", "mbasic.", 1).replace("www.", "mbasic.")
            self.navegador.navegar(url + "?v=timeline")
        else: 
            self.navegador.navegar(LINKS['base_facebook'] + usuario_pesquisa + "?v=timeline")
        sleep(3)

        # Verifica se o form login está na tela
        if "login.php" in self.navegador.url_atual():
            self.navegador.encerrar_navegador()
            sleep(2)
            self.navegador = Navegador(salvar_cache=PATH['arquivo_cache'], tela_cheia=True)
            self.login_facebook(CREDENCIAIS['email_facebook'], CREDENCIAIS['senha_facebook'])
            sleep(3)
            self.navegador.encerrar_navegador()
            sleep(2)
            self.navegador = Navegador(somente_terminal=True, salvar_cache=PATH['arquivo_cache'], tela_cheia=True)
            if self.e_url(url):
                if "mbasic" not in url:
                    url = url.replace("web.", "mbasic.", 1).replace("www.", "mbasic.")
                self.navegador.navegar(url + "?v=timeline")
            else: 
                self.navegador.navegar(LINKS['base_facebook'] + usuario_pesquisa + "?v=timeline")
            sleep(3)

        self.navegador.esperar('tag_name', 'article')
        self.navegador.esperar('xpath', XPATHS['ver_mais_stories'])

        loop = 1
        parar = 0
        print("Coletando dados do perfil:", usuario_pesquisa)
        # Enquanto tiver btn ver_mais_stories, continua..
        while True:
            url_atual = self.navegador.url_atual()
            data_atual = datetime.now().strftime("%d.%m.%Y")
            elementos_postagens = self.navegador.obter_elementos('TAG_NAME', 'article')
            for i, article in enumerate(elementos_postagens):
                nome_arquivos_imagens = ""
                nome_arquivos_video = ""
                texto_postagem = ""

                self.navegador.esperar(By.XPATH, f'//article[{i+1}]//a[text()="História completa"]')
                url_postagem = article.find_element(By.XPATH, f'//article[{i+1}]//a[text()="História completa"]').get_attribute("href").replace("mbasic", "web")
                data = article.find_element(By.XPATH, f'//article[{i+1}]//abbr').text
                try:
                    texto_postagem = article.find_element(By.XPATH, f'//article[{i+1}]//p').text
                except:
                    texto_postagem = ""

                if data[-1].strip().lower() == 'm' or data[-1].strip().lower() == 'h':
                    data_postagem = datetime.now().strftime("%d/%m/%Y")

                elif data.split(" ")[0].strip().lower() == 'ontem':
                    data_postagem = datetime.now() - timedelta(days=dias)

                elif data[-1].strip().lower() == 'd':
                    # Dias atrás
                    dias = data[0]
                    data_postagem = datetime.now() - timedelta(days=dias)
                else:
                    dia = data.split(" ")[0].zfill(2)
                    mes = str(self.converte_nome_mes(data.split(" ")[2])).zfill(2)

                    if ":" in data.split(" ")[4]:
                        ano = datetime.now().year
                    else:
                        ano = data.split(" ")[4]
                    data_postagem = f"{dia}/{mes}/{ano}"

                if filtrar_periodo:
                    # Verifica se a data está dentro do período
                    data_inicio = datetime.strptime(periodo_inicio, "%d/%m/%Y")
                    data_verificar = datetime.strptime(data_postagem, "%d/%m/%Y")
                    data_final = datetime.strptime(periodo_final, "%d/%m/%Y")

                    if not data_inicio <= data_verificar <= data_final:
                        parar = 1
                        break

                # Verifica se o arquivo de OUTPUT está criado
                nome_arquivo = f"\{usuario_pesquisa}\{data_atual}"
                if not os.path.exists(PATH['saida_facebook'] + f'\{usuario_pesquisa}'):
                    os.mkdir(PATH['saida_facebook'] + f'\{usuario_pesquisa}')
                if not os.path.exists(PATH['saida_facebook'] + f'\{nome_arquivo}'):
                    os.mkdir(PATH['saida_facebook'] + f'\{nome_arquivo}')
                
                # Verifica se tem imagem
                if len(article.find_elements(By.XPATH, f'//article[{i+1}]//a/img')) >= 1:
                    imagens = article.find_elements(By.XPATH, f'//article[{i+1}]//a/img')
                    print(f"Salvando Imagem do post #{i+1}. Página: {loop}")
                    for img in imagens:
                        imagem_url = img.get_attribute("src").replace("mbasic", "web")
                        hora = str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond)
                        if not os.path.exists(PATH['saida_facebook'] + nome_arquivo + f"\img"):
                            os.makedirs(PATH['saida_facebook'] + nome_arquivo + f"\img")
                        self.baixar_conteudo(imagem_url, PATH['saida_facebook'] + nome_arquivo + f"\img\{hora}.png")
                        nome_arquivos_imagens = nome_arquivos_imagens + nome_arquivo + f"\img\{hora}.png, "
                
                # Verifica se tem vídeo
                if len(article.find_elements(By.XPATH, f'//article[{i+1}]//a[@aria-label="Assistir ao vídeo"]')) > 0:
                    print(f"Salvando Vídeo do post #{i+1}. Página: {loop}")
                    video_url = article.find_element(By.XPATH, f'//article[{i+1}]//a[@aria-label="Assistir ao vídeo"]').get_attribute("href")

                    janela = self.navegador.id_janela_atual()
                    self.navegador.abrir_nova_janela()
                    self.navegador.navegar(video_url)
                    while self.navegador.url_atual() == video_url:
                        pass
                    url = self.navegador.url_atual()
                    self.navegador.fechar_janela()
                    self.navegador.mudar_janela(janela)

                    hora = str(datetime.now().hour) + str(datetime.now().minute) + str(datetime.now().second) + str(datetime.now().microsecond)
                    if not os.path.exists(PATH['saida_facebook'] + nome_arquivo + f"\movie"):
                        os.makedirs(PATH['saida_facebook'] + nome_arquivo + f"\movie")
                    self.baixar_conteudo(url, PATH['saida_facebook'] + nome_arquivo + f"\movie\{hora}.mp4")
                    nome_arquivos_video = nome_arquivos_video + nome_arquivo + f"\movie\{hora}.mp4, "

                # Salva os dados na planilha
                print(f"Gravando dados do post #{i+1}. Página: {loop}")
                self.adicionar_dados_excel(PATH['saida_facebook'] + nome_arquivo + f"\dados.xlsx",
                                           data_postagem, 
                                           url_postagem, 
                                           texto_postagem, 
                                           nome_arquivos_imagens, 
                                           nome_arquivos_video
                                           )
            if parar == 1:
                print("Salvando dados no Excel.")
                sleep(1)
                print("Scraping realizado com sucesso!")
                break
            if len(self.navegador.obter_elementos('XPATH', XPATHS['ver_mais_stories'])) > 0:
                self.navegador.esperar('XPATH', XPATHS['ver_mais_stories'])
                proximo = self.navegador.obter_elemento('XPATH', XPATHS['ver_mais_stories']).get_attribute("href")
                self.navegador.navegar(proximo)
                while self.navegador.url_atual == url_atual:
                    pass
                self.navegador.esperar('tag_name', 'article')
            else:
                print("Salvando dados no Excel.")
                sleep(1)
                print("Scraping realizado com sucesso!")
                break
            loop += 1

    def carregar(self):
        chars = "/—\|" 
        for _ in range(2):
            for char in chars:
                print(f"\033[34m\rProcessando {char}\033[0m", end="")
                sleep(0.1)

    def carregar_iniciando(self):
        chars = ['.', '..', '...']
        for _ in range(2):
            for char in chars:
                print(f"\033[34m\rIniciando Raspagem {char}\033[0m", end="")
                sleep(0.3)

    def e_url(self, texto=str):
        """Verifica se a string é um Link de URL."""

        if "http://" or "www." or "https://" in texto:
            return True
        else:
            return False
        
    def login_facebook(self, usuario=str, senha=str):
        """Faz login no Facebook."""
        self.navegador.navegar(LINKS['login_facebook'])
        self.navegador.esperar('XPATH', XPATHS['email'])
        self.navegador.esperar('XPATH', XPATHS['senha'])
        input_email = self.navegador.obter_elemento('XPATH', XPATHS['email'])
        input_senha = self.navegador.obter_elemento('XPATH', XPATHS['senha'])

        self.navegador.clicar_elemento(input_email)
        self.navegador.enviar_teclas(usuario, espera_entre_as_teclas=0.2)
        self.navegador.espera_aleatoria(1, 2)

        self.navegador.clicar_elemento(input_senha)
        self.navegador.enviar_teclas(senha)
        self.navegador.espera_aleatoria(1, 2)

        input_senha.send_keys(Keys.ENTER)

    def baixar_conteudo(self, link, nome_arquivo):
        try:
            response = requests.get(link)
            if response.status_code == 200:
                with open(nome_arquivo, 'wb') as arquivo:
                    arquivo.write(response.content)
            else:
                print(f'Erro. Status code: {response.status_code}')
        except Exception as e:
            print(f'Ocorreu um erro: {str(e)}')

    def converte_nome_mes(self, nome_mes):
        meses = {
            'janeiro': 1,
            'fevereiro': 2,
            'março': 3,
            'abril': 4,
            'maio': 5,
            'junho': 6,
            'julho': 7,
            'agosto': 8,
            'setembro': 9,
            'outubro': 10,
            'novembro': 11,
            'dezembro': 12
        }
        
        # Converte o nome do mês para minúsculas e verifica se está no dicionário
        nome_mes = nome_mes.lower()
        if nome_mes in meses:
            return meses[nome_mes]
        else:
            return None 
    
    def adicionar_dados_excel(self, nome_planilha, data, url, texto, imagens, videos):
        if not os.path.exists(nome_planilha):
            wb = Workbook()
            ws = wb.active
            ws.append(["Data da Postagem", "URL da Postagem", "Texto da Postagem", "Nome dos Arquivos das Imagens", "Nome dos Arquivos dos Vídeos"])
        else:
            wb = load_workbook(nome_planilha)
            ws = wb.active
        
        ws.append([data, url, texto, imagens, videos])
        wb.save(nome_planilha)